import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime
import re
import anthropic
import tkinter as tk
from tkinter import filedialog, messagebox
import json
import os
import requests

# Anthropic API Configuration
ANTHROPIC_API_KEY = "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"

# JSONBin Kill Switch API Configuration
JSONBIN_API_URL = "https://api.jsonbin.io/v3/b/688926957b4b8670d8a95cd3/latest"
JSONBIN_API_KEY = "$2a$10$78AAni5Tx5vba9VOjwQZjeEh6.0UFPZ4155iYE.RmDCyfEm6sxyWm"  # Replace with your actual JSONBin API key

# Settings file to remember last locations
SETTINGS_FILE = "accrual_settings.json"

def load_settings():
    """Load saved settings for file locations"""
    default_settings = {
        "last_checkbook_dir": os.getcwd(),
        "last_output_dir": os.getcwd(),
        "last_run_time": None,
        "last_output_file": None
    }
    
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
                # Merge with defaults to handle missing keys
                for key, value in default_settings.items():
                    if key not in settings:
                        settings[key] = value
                return settings
    except Exception as e:
        print(f"Warning: Could not load settings: {e}")
    
    return default_settings

def save_settings(settings):
    """Save settings for next run"""
    try:
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(settings, f, indent=2)
    except Exception as e:
        print(f"Warning: Could not save settings: {e}")

def check_license():
    """
    TRUE remote kill switch - app MUST reach server to work
    No offline fallback - if server is unreachable, app is blocked
    """
    import uuid
    import hashlib
    
    try:
        print("🔍 Contacting license server (required)...")
        
        # Generate hardware ID for tracking (optional)
        try:
            mac = hex(uuid.getnode())
            hardware_id = hashlib.sha256(mac.encode()).hexdigest()[:16]
        except:
            hardware_id = "unknown"
        
        response = requests.get(
            JSONBIN_API_URL,
            headers={'X-Master-Key': JSONBIN_API_KEY},
            timeout=15  # Wait up to 15 seconds
        )
        
        if response.status_code == 200:
            data = response.json()
            enabled = data.get('record', {}).get('enabled', False)
            
            if enabled:
                print("✅ License server says: ENABLED")
                return True
            else:
                messagebox.showerror("Access Denied", 
                                   "Service has been disabled by administrator.")
                return False
        else:
            messagebox.showerror("Connection Error", 
                               f"Cannot reach license server (HTTP {response.status_code}). "
                               "Please check your internet connection and try again.")
            return False
            
    except requests.exceptions.Timeout:
        messagebox.showerror("Connection Timeout", 
                           "License server is taking too long to respond. "
                           "Please check your internet connection and try again.")
        return False
        
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Network Error", 
                           "Cannot reach license server. Please check your internet connection "
                           "and try again.\n\nIf the problem persists, contact administrator.")
        return False

def get_cutoff_month():
    """
    Determine which months to include in accrual based on current date and day
    """
    now = datetime.now()
    current_month = now.month
    current_day = now.day
    
    # If before the 15th, don't include current month
    if current_day < 15:
        cutoff_month = current_month - 1
    else:
        cutoff_month = current_month
    
    # Handle January edge case
    if cutoff_month <= 0:
        cutoff_month = 12
    
    month_names = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                   'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
    
    # Get months to include (up to cutoff month)
    months_to_include = month_names[:cutoff_month]
    
    print(f"📅 Current date: {now.strftime('%B %d, %Y')}")
    print(f"📋 Including months through: {month_names[cutoff_month-1]}")
    
    return months_to_include, cutoff_month

def create_smart_description(original_description, po_number, invoice, max_length=40):
    """
    Use Anthropic AI to create clean, standardized descriptions for ALL entries
    Format: Invoice number (if exists) OR clean description (if no invoice)
    """
    if not original_description or original_description == 'nan' or str(original_description).strip() == '':
        original_description = "Service/Supply"
    
    original_description = str(original_description).strip()
    po_number = str(po_number).strip() if po_number else ""
    invoice = str(invoice).strip() if invoice and invoice != 'nan' else ""
    
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        # Different prompts based on whether we have an invoice
        if invoice and invoice != "":
            # If we have an invoice, just return the invoice number
            return f"INV{invoice}"
            
        else:
            # If no invoice, create clean description (NO PO number - that's added separately)
            prompt = f"""Create a clean accounting description for what was purchased or what service was provided.

Original Description: {original_description}

Extract the most important information about the actual item/service. Remove redundant PO references, vendor names (that's in another field), and unnecessary prefixes like "ENG/", "INN/", "MH/". Focus ONLY on the actual service or item purchased. Keep it under {max_length} characters.

Examples:
- "ENG/PO25377 Republic Services/Waste Removal June Monthly Service" → "Waste Removal Monthly Service"
- "PO24913 ATS Inland Monthly Service Contract 2025" → "Monthly Service Contract"
- "ENG/PO#25831 NALCO Water Treatment Chemicals" → "Water Treatment Chemicals"
- "Emergency Uniforms for Staff" → "Emergency Uniforms"

Return only the clean description, no quotes, no PO numbers."""

            response = client.messages.create(
                model="claude-3-haiku-20240307",  # Cheapest model
                max_tokens=60,
                temperature=0,
                messages=[{
                    "role": "user", 
                    "content": prompt
                }]
            )
            
            clean_description = response.content[0].text.strip().strip('"\'')
            
            # Ensure it's within max_length
            if len(clean_description) > max_length:
                clean_description = clean_description[:max_length-3] + "..."
                
            return clean_description
        
    except Exception as e:
        print(f"Warning: AI description creation failed for PO {po_number}: {str(e)}")
        
        # Fallback: create description manually
        if invoice and invoice != "":
            return f"INV{invoice}"
        else:
            # Try to extract key words from original description
            clean_desc = original_description
            # Remove common prefixes
            clean_desc = re.sub(r'^(ENG?|INN?|MH)/PO\d+\s*[^/]*/', '', clean_desc, flags=re.IGNORECASE)
            clean_desc = re.sub(r'^PO\d+\s*', '', clean_desc, flags=re.IGNORECASE)
            clean_desc = clean_desc.strip()
            
            if len(clean_desc) > max_length:
                clean_desc = clean_desc[:max_length-3] + "..."
                
            return clean_desc or "Service/Supply"

def extract_accruals_from_checkbook(checkbook_file):
    """
    Extract all uncleared transactions from the checkbook Excel file
    """
    print(f"Reading checkbook file: {checkbook_file}")
    
    # Get date cutoff logic
    months_to_include, cutoff_month_num = get_cutoff_month()
    
    # Read all sheets from the workbook
    xl_file = pd.ExcelFile(checkbook_file)
    
    # Sheets to exclude
    excluded_sheets = ['Summary', 'Non-Eng ACTIVE', 'Non-Eng CLEARED']
    
    # Store all accrual transactions
    accruals = []
    
    for sheet_name in xl_file.sheet_names:
        if sheet_name in excluded_sheets:
            print(f"Skipping sheet: {sheet_name}")
            continue
            
        print(f"Processing sheet: {sheet_name}")
        
        try:
            # Read the sheet
            df = pd.read_excel(checkbook_file, sheet_name=sheet_name, header=None)
            
            # Extract GL account and description from first row
            gl_description = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else "Unknown"
            gl_account = str(df.iloc[0, 1]) if not pd.isna(df.iloc[0, 1]) else "Unknown"
            
            # All month names for reference
            all_months = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
                         'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']
            
            # Track current month context while processing
            current_month_context = None
            
            # Process each row looking for transactions
            for idx, row in df.iterrows():
                # Skip if not enough columns or empty row
                if len(row) < 9 or pd.isna(row.iloc[0]):
                    continue
                
                # Check if this looks like a transaction row (has vendor and amount)
                vendor_cell = str(row.iloc[0]).strip()
                
                # Check if this row is a month header
                if vendor_cell.upper() in all_months:
                    current_month_context = vendor_cell.upper()
                    print(f"  Found month section: {current_month_context}")
                    continue
                
                # Skip other header rows, total rows, balance rows
                if (vendor_cell.upper() in ['VENDOR', 'TOTAL', 'BALANCE'] or
                    vendor_cell == '' or vendor_cell == 'nan'):
                    continue
                
                # Skip transactions from months beyond our cutoff
                if current_month_context and current_month_context not in months_to_include:
                    continue
                
                # Check if second column has an amount (numeric)
                try:
                    amount = pd.to_numeric(row.iloc[1])
                    if pd.isna(amount) or amount == 0:
                        continue
                except (ValueError, TypeError):
                    continue
                
                # Check if PO number exists (REQUIRED)
                po_number = str(row.iloc[2]) if not pd.isna(row.iloc[2]) else ""
                if not po_number or po_number.strip() == "" or po_number == "nan":
                    continue  # Skip if no PO number
                
                # Extract transaction details
                vendor = vendor_cell
                # Remove parenthetical amounts from vendor name as requested
                vendor = re.sub(r'\s*\([^)]*\)', '', vendor).strip()
                
                po_date = row.iloc[3] if not pd.isna(row.iloc[3]) else ""
                delivery_date = row.iloc[4] if not pd.isna(row.iloc[4]) else ""
                received_date = row.iloc[5] if not pd.isna(row.iloc[5]) else ""
                invoice = str(row.iloc[6]) if not pd.isna(row.iloc[6]) else ""
                cleared = str(row.iloc[7]) if not pd.isna(row.iloc[7]) else ""
                description = str(row.iloc[8]) if not pd.isna(row.iloc[8]) else ""
                
                # Create smart, standardized description using AI
                smart_description = create_smart_description(description, po_number, invoice)
                
                # Check if transaction should be accrued (Cleared ≠ "Y")
                if cleared.upper().strip() != 'Y':
                    # Create PO/Invoice reference
                    ref_po_invoice = po_number
                    if invoice and invoice != 'nan' and invoice.strip():
                        ref_po_invoice += f" {invoice}"
                    
                    # Create JE Entry description (use the smart description)
                    je_entry = f"{po_number} {smart_description}".strip()
                    je_entry_30 = je_entry[:30] + "..." if len(je_entry) > 30 else je_entry
                    
                    accrual = {
                        'GL_Account': gl_account,
                        'Amount': amount,
                        'Vendor': vendor,
                        'Ref_PO_Invoice': ref_po_invoice,
                        'Description': smart_description,
                        'JE_Entry': je_entry,
                        'JE_Entry_30': je_entry_30,
                        'PO_Date': po_date,
                        'Delivery_Date': delivery_date,
                        'Received_Date': received_date,
                        'Invoice': invoice,
                        'Sheet_Name': sheet_name,
                        'Month_Context': current_month_context
                    }
                    
                    accruals.append(accrual)
                    print(f"    Added accrual: {vendor} - ${amount} ({current_month_context})")
                    
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {str(e)}")
            continue
    
    return accruals

def create_accrual_report(accruals, output_file):
    """
    Create a beautifully styled Excel accrual report matching the template format
    """
    print(f"Creating styled accrual report with {len(accruals)} transactions")
    
    # Convert to DataFrame
    df = pd.DataFrame(accruals)
    
    if df.empty:
        print("No accruals found!")
        return
    
    # Sort by GL Account, then by Vendor (as requested)
    df = df.sort_values(['GL_Account', 'Vendor', 'PO_Date'])
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accruals"
    
    # Define colors and styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # White text
    
    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    
    # Alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    
    # Total row styling
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
    total_font = Font(name="Calibri", size=11, bold=True)
    
    # Add headers matching template format
    headers = [
        'GL Account #',
        'Amount', 
        'Vendor',
        '',  # Empty column D (will be hidden)
        'Ref PO / Invoice',
        '',  # Empty column F (will be hidden)
        'Description',
        'JE Entry',  # Will be hidden
        'JE Entry (30 chars)'
    ]
    
    # Write and style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )
    
    # Write data rows with styling
    row_num = 2
    total_amount = 0
    current_gl = None
    
    for idx, (_, accrual) in enumerate(df.iterrows()):
        # Determine row fill (alternating colors, but group by GL account)
        if current_gl != accrual['GL_Account']:
            current_gl = accrual['GL_Account']
            gl_row_count = 0
        
        row_fill = light_fill if gl_row_count % 2 == 0 else white_fill
        gl_row_count += 1
        
        # Write data
        cells_data = [
            (accrual['GL_Account'], data_alignment),
            (accrual['Amount'], number_alignment), 
            (accrual['Vendor'], data_alignment),
            ('', data_alignment),  # Empty column D
            (accrual['Ref_PO_Invoice'], data_alignment),
            ('', data_alignment),  # Empty column F
            (accrual['Description'], data_alignment),
            (accrual['JE_Entry'], data_alignment),
            (accrual['JE_Entry_30'], data_alignment)
        ]
        
        for col, (value, alignment) in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = alignment
            cell.border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Format amount as currency
            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00'
        
        total_amount += accrual['Amount']
        row_num += 1
    
    # Add total row with special styling
    total_cells_data = [
        ("TOTAL", data_alignment),
        (total_amount, number_alignment),
        ('', data_alignment),
        ('', data_alignment),
        ('', data_alignment),
        ('', data_alignment),
        ('', data_alignment),
        ('', data_alignment),
        ('', data_alignment)
    ]
    
    for col, (value, alignment) in enumerate(total_cells_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = alignment
        cell.border = Border(
            left=Side(style='medium', color='4472C4'),
            right=Side(style='medium', color='4472C4'),
            top=Side(style='medium', color='4472C4'),
            bottom=Side(style='medium', color='4472C4')
        )
        
        # Format total amount as currency
        if col == 2:
            cell.number_format = '$#,##0.00'
    
    # Adjust column widths for optimal readability
    column_widths = [
        ('A', 15),  # GL Account
        ('B', 12),  # Amount
        ('C', 25),  # Vendor
        ('D', 3),   # Hidden column
        ('E', 30),  # Ref PO/Invoice
        ('F', 3),   # Hidden column
        ('G', 35),  # Description
        ('H', 35),  # JE Entry (hidden)
        ('I', 35)   # JE Entry 30 chars
    ]
    
    for col_letter, width in column_widths:
        ws.column_dimensions[col_letter].width = width
    
    # Freeze the header row for easier scrolling
    ws.freeze_panes = 'A2'
    
    # Hide columns D, F, and H as requested
    ws.column_dimensions['D'].hidden = True
    ws.column_dimensions['F'].hidden = True  
    ws.column_dimensions['H'].hidden = True
    
    # Add auto-filter to headers
    ws.auto_filter.ref = f"A1:I{row_num}"
    
    # Set print settings for professional output
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Set margins
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Save the file
    wb.save(output_file)
    print(f"📊 Beautifully styled accrual report saved to: {output_file}")
    print(f"✅ Features: Alternating rows, frozen headers, auto-filter, currency formatting")
    print(f"✅ Columns D, F, and H are hidden")
    print(f"💰 Total accrual amount: ${total_amount:,.2f}")
    
    # Print summary by GL account
    gl_summary = df.groupby('GL_Account')['Amount'].sum().sort_values(ascending=False)
    print(f"\n📋 Accrual Summary by GL Account:")
    for gl, amount in gl_summary.items():
        print(f"  {gl}: ${amount:,.2f}")

def select_files():
    """
    Show file dialogs to select checkbook and output location
    """
    settings = load_settings()
    
    # Hide the root tkinter window
    root = tk.Tk()
    root.withdraw()
    
    # Show welcome message
    messagebox.showinfo("Accrual Generator", 
                       "Welcome to the Accrual Report Generator!\n\n"
                       "First, select your checkbook Excel file.\n"
                       "Then choose where to save the accrual report.")
    
    # Select checkbook file
    checkbook_file = filedialog.askopenfilename(
        title="Select Checkbook Excel File",
        initialdir=settings["last_checkbook_dir"],
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )
    
    if not checkbook_file:
        messagebox.showinfo("Cancelled", "No checkbook file selected. Exiting.")
        return None, None
    
    # Update checkbook directory
    settings["last_checkbook_dir"] = os.path.dirname(checkbook_file)
    
    # Generate output filename with better format
    now = datetime.now()
    date_str = now.strftime("%m_%d_%Y")
    time_str = now.strftime("%I_%M%p").lower()
    output_filename = f"Accrual_Prelim_{date_str}_{time_str}.xlsx"
    
    # Select output location
    output_file = filedialog.asksaveasfilename(
        title="Save Accrual Report As",
        initialdir=settings["last_output_dir"],
        initialfile=output_filename,
        defaultextension=".xlsx",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )
    
    if not output_file:
        messagebox.showinfo("Cancelled", "No output location selected. Exiting.")
        return None, None
    
    # Update settings
    settings["last_output_dir"] = os.path.dirname(output_file)
    settings["last_run_time"] = now.isoformat()
    settings["last_output_file"] = output_file
    save_settings(settings)
    
    # Show last run info if available
    if settings.get("last_run_time"):
        try:
            last_run = datetime.fromisoformat(settings["last_run_time"])
            messagebox.showinfo("Previous Run", 
                              f"Last accrual generated:\n"
                              f"Date: {last_run.strftime('%B %d, %Y at %I:%M %p')}\n"
                              f"File: {os.path.basename(settings.get('last_output_file', 'Unknown'))}")
        except:
            pass
    
    root.destroy()
    return checkbook_file, output_file

def main():
    """
    Main function to run the accrual maker
    """
    print("🚀 Accrual Report Generator Starting...")
    print("=" * 50)
    
    # Check license first
    if not check_license():
        return
    
    # Select files using GUI
    checkbook_file, output_file = select_files()
    if not checkbook_file or not output_file:
        return
    
    print(f"📁 Checkbook file: {os.path.basename(checkbook_file)}")
    print(f"💾 Output file: {os.path.basename(output_file)}")
    print("=" * 50)
    
    print("🤖 Using Anthropic Claude 3 Haiku for ALL description processing...")
    print("📝 AI creating standardized descriptions: Invoice OR CleanDescription")
    print("📅 Smart date filtering: Only accruing appropriate months based on current date")
    
    try:
        # Extract accruals from checkbook
        accruals = extract_accruals_from_checkbook(checkbook_file)
        
        if not accruals:
            message = ("No uncleared transactions found for accrual.\n\n"
                      "Requirements:\n"
                      "• Vendor name present\n"
                      "• Amount > 0\n"
                      "• PO number present\n" 
                      "• Cleared ≠ 'Y'\n"
                      "• Within appropriate date range")
            print(message)
            messagebox.showwarning("No Accruals Found", message)
            return
        
        # Create accrual report
        create_accrual_report(accruals, output_file)
        
        # Success messages
        success_msg = (f"✅ Accrual report completed successfully!\n\n"
                      f"📄 File: {os.path.basename(output_file)}\n"
                      f"📋 Transactions: {len(accruals)}\n"
                      f"🤖 AI-enhanced descriptions\n"
                      f"📅 Date-aware filtering applied")
        
        print("\n" + "=" * 50)
        print(success_msg.replace("✅ ", "").replace("\n\n", "\n"))
        print("=" * 50)
        
        messagebox.showinfo("Success!", success_msg)
        
        # Ask if user wants to open the file
        if messagebox.askyesno("Open File?", "Would you like to open the accrual report now?"):
            try:
                os.startfile(output_file)  # Windows
            except:
                try:
                    os.system(f'open "{output_file}"')  # Mac
                except:
                    print(f"Please manually open: {output_file}")
        
    except FileNotFoundError:
        error_msg = f"❌ Error: Could not find the selected checkbook file."
        print(error_msg)
        messagebox.showerror("File Not Found", error_msg)
    except Exception as e:
        error_msg = f"❌ Error: {str(e)}"
        print(error_msg)
        messagebox.showerror("Error", f"An error occurred:\n\n{str(e)}")

# Installation requirements:
# pip install pandas openpyxl anthropic requests

if __name__ == "__main__":
    main()